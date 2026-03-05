#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""M5: Build master tables (job-level + aggregated summary).

Outputs (by default under --out-dir):
- master_table.csv           (one row per job)
- master_summary.csv         (grouped by run_id + baseline)
- master_table.xlsx          (sheet: master_table, summary)

Design goals (per your latest spec):
- C0: output as missing_imports/total_imports (aggregated in summary)
  - total_imports prefers a stable baseline computed from git-tracked Python files
    (build_output/pyright/baseline_imports.json), falling back to the per-job pyright metric.
- C1/C3/C4: output as success_count/total_applicable_repos, where applicability is defined by
  scripts_repos_test_categories.csv (snapshotted into the run dir when available).
- C2: output as success_count/job_count (CUDA availability check is always applicable).
- C5: keep hallucination counts as integers (sum in summary)
- agent_wall_time_avg_sec / env_size_avg_gb: averaged over valid jobs

This script is *best-effort*:
- If a stage results.json is missing, that job is counted as "failed" for that stage in summary.
- If C0 total imports is missing or 0, c0_missing_over_total is left blank in summary.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from m5.utils import (
    normalize_status,
    parse_nvidia_smi_text,
    parse_repo_full_name,
    read_json,
    safe_get,
    try_read_text,
    write_json,
)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


C0_REPO_BASELINE_TOTALS_PATH = Path(__file__).resolve().with_name("c0_repo_baseline_totals.json")


def _normalize_repo_key(v: Any) -> str:
    return str(v or "").strip().lower()


def _load_c0_repo_baseline_totals() -> Dict[str, int]:
    """Load canonical repo-level C0 totals so denominator stays stable per repo."""
    obj = read_json(C0_REPO_BASELINE_TOTALS_PATH)
    if not isinstance(obj, dict):
        return {}

    repos = obj.get("repos")
    if not isinstance(repos, dict):
        return {}

    out: Dict[str, int] = {}
    for k, v in repos.items():
        nk = _normalize_repo_key(k)
        iv: Optional[int] = None
        if isinstance(v, int):
            iv = v
        elif isinstance(v, float):
            iv = int(v)
        else:
            try:
                iv = int(float(str(v).strip()))
            except Exception:
                iv = None
        if nk and iv is not None and iv > 0:
            out[nk] = iv
    return out


C0_REPO_BASELINE_TOTALS = _load_c0_repo_baseline_totals()


# -----------------------------
# Job-level columns (one row per job)
# -----------------------------
JOB_COLUMNS = [
    # identity
    "run_id",
    "job_id",
    "repo_full_name",
    "repo_url",
    "commit_sha",
    "baseline",
    "hardware_bucket",
    "script_id",
    # applicability (from scripts_repos_test_categories.csv; filled in main)
    "supports_cpu",
    "supports_single_gpu",
    "supports_multi_gpu",
    # C0~C5 (raw per-job)
    "c0_missing_imports",
    "c0_total_imports",
    "c0_total_imports_baseline",
    "c1_cpu_status",
    "c2_cuda_status",
    "c3_single_gpu_status",
    "c4_multi_gpu_status",
    "c5_path_hallucinations_count",
    "c5_version_hallucinations_count",
    "c5_capability_hallucinations_count",
    # process metrics
    "agent_wall_time_sec",
    "env_prefix_size_mb",
    # machine/image
    "image_id",
    "gpu_name",
    "driver_version",
    "cuda_runtime",
    # debug
    "job_dir",
]


# -----------------------------
# Summary columns (grouped by run_id + baseline)
# -----------------------------
SUMMARY_COLUMNS = [
    "run_id",
    "baseline",
    "job_count",
    "c0_missing_sum",
    "c0_total_sum",
    "c0_total_sum_reported",
    "c0_missing_over_total",
    "c1_cpu_denom",
    "c2_cuda_denom",
    "c3_single_gpu_denom",
    "c4_multi_gpu_denom",
    "c1_cpu_success_over_all",
    "c2_cuda_success_over_all",
    "c3_single_gpu_success_over_all",
    "c4_multi_gpu_success_over_all",
    "c1_cpu_success",
    "c1_cpu_failed",
    "c1_cpu_skipped",
    "c2_cuda_success",
    "c2_cuda_failed",
    "c2_cuda_skipped",
    "c3_single_gpu_success",
    "c3_single_gpu_failed",
    "c3_single_gpu_skipped",
    "c4_multi_gpu_success",
    "c4_multi_gpu_failed",
    "c4_multi_gpu_skipped",
    "c5_path_sum",
    "c5_version_sum",
    "c5_capability_sum",
    "c5_total_sum",
    "agent_wall_time_avg_sec",
    "agent_time_valid_jobs",
    "env_size_avg_gb",
    "env_valid_jobs",
    # Langfuse token metrics (optional; exported by tools/langfuse/export_run_tokens.py)
    "agent_tokens_input_sum",
    "agent_tokens_output_sum",
    "agent_tokens_total_sum",
    "agent_tokens_total_avg",
]


# -----------------------------
# Helpers
# -----------------------------
def find_job_dirs(results_root: Path) -> List[Path]:
    # Job dir = contains job_summary.json
    return sorted({p.parent for p in results_root.rglob("job_summary.json")})


def load_stage_results(build_output_dir: Path, stage: str) -> Optional[Dict[str, Any]]:
    return read_json(build_output_dir / stage / "results.json")


def coerce_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        return int(x)
    s = str(x).strip()
    try:
        return int(float(s))
    except Exception:
        return None


def coerce_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    try:
        return float(s)
    except Exception:
        return None


def load_langfuse_run_token_sums(run_dir: Path) -> Dict[str, Optional[int]]:
    """Load aggregated token sums for a run from <run_dir>/langfuse_tokens.json.

    The file is produced by tools/langfuse/export_run_tokens.py.
    Returns a dict with keys: input_tokens, output_tokens, total_tokens.
    Values may be None if missing/unavailable.
    """
    p = run_dir / 'langfuse_tokens.json'
    if not p.exists():
        return {
            'input_tokens': None,
            'output_tokens': None,
            'total_tokens': None,
        }

    obj = read_json(p)
    if not isinstance(obj, dict):
        return {
            'input_tokens': None,
            'output_tokens': None,
            'total_tokens': None,
        }

    metrics = obj.get('metrics')
    if not isinstance(metrics, dict):
        return {
            'input_tokens': None,
            'output_tokens': None,
            'total_tokens': None,
        }

    in_tok = coerce_int(metrics.get('input_tokens'))
    out_tok = coerce_int(metrics.get('output_tokens'))
    tot_tok = coerce_int(metrics.get('total_tokens'))

    # Derive total if only input/output are present
    if tot_tok is None and in_tok is not None and out_tok is not None:
        tot_tok = in_tok + out_tok

    return {
        'input_tokens': in_tok,
        'output_tokens': out_tok,
        'total_tokens': tot_tok,
    }


def _normalize_repo_url(url: str) -> str:
    s = (url or "").strip()
    if s.endswith(".git"):
        s = s[:-4]
    return s.rstrip("/")


def _normalize_yes(v: Any) -> bool:
    return str(v or "").strip().lower() in ("yes", "y", "true", "1")


def _is_supported(v: Any) -> bool:
    """Interpret supports_* flags. Default to supported when unknown/blank."""
    s = str(v or "").strip().lower()
    if s in ("no", "n", "false", "0"):
        return False
    return True


def c0_reference_total_for_repo(repo_full_name: str) -> Optional[int]:
    return C0_REPO_BASELINE_TOTALS.get(_normalize_repo_key(repo_full_name))


def resolve_categories_csv(out_dir: Path) -> Optional[Path]:
    """Resolve scripts_repos_test_categories.csv with best-effort precedence.

    - Prefer a snapshot saved under the run dir (created by host_orchestrator.py)
    - Fallback to a copy under out_dir
    - Fallback to the harness root (CWD when invoked by host_orchestrator.py)
    """
    candidates = [
        out_dir / "scripts_repos_test_categories.snapshot.csv",
        out_dir / "scripts_repos_test_categories.csv",
        Path.cwd() / "scripts_repos_test_categories.csv",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def load_categories_index(path: Path) -> Dict[str, Dict[str, Dict[str, str]]]:
    by_slug: Dict[str, Dict[str, str]] = {}
    by_repo: Dict[str, Dict[str, str]] = {}
    by_git: Dict[str, Dict[str, str]] = {}

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            r = {k: (v or "").strip() for k, v in row.items()}
            slug = r.get("repo_slug") or ""
            repo = r.get("repo") or ""
            git_link = _normalize_repo_url(r.get("git_link") or "")
            if slug:
                by_slug[slug] = r
            if repo:
                by_repo[repo] = r
            if git_link:
                by_git[git_link] = r

    return {"by_slug": by_slug, "by_repo": by_repo, "by_git": by_git}


def find_categories_row(
    idx: Dict[str, Dict[str, Dict[str, str]]],
    script_id: str,
    repo_url: str,
    repo_full_name: str,
) -> Optional[Dict[str, str]]:
    sid = (script_id or "").strip()
    if sid and sid in idx.get("by_slug", {}):
        return idx["by_slug"][sid]
    rfn = (repo_full_name or "").strip()
    if rfn and rfn in idx.get("by_repo", {}):
        return idx["by_repo"][rfn]
    url = _normalize_repo_url(repo_url or "")
    if url and url in idx.get("by_git", {}):
        return idx["by_git"][url]
    return None


def top_level_import_name(x: Any) -> Optional[str]:
    """Normalize an import/module string to a top-level package name.

    Examples:
      - "a.b.c" -> "a"
      - "a" -> "a"
      - "No module named 'a.b'" -> "a"
      - ".relative" -> None
    """
    if not isinstance(x, str):
        return None
    s = x.strip()
    if not s:
        return None

    # Handle common error strings like: "No module named 'a.b'"
    if "no module named" in s.lower() and ("'" in s or '"' in s):
        for q in ("'", '"'):
            parts = s.split(q)
            if len(parts) >= 2 and parts[1].strip():
                s = parts[1].strip()
                break

    s = s.strip().strip("\"'").lstrip(".")
    if not s:
        return None

    # Keep first token and remove trailing punctuation
    s = s.split()[0].rstrip(",:;)")
    if not s:
        return None

    return s.split(".")[0] or None


def unique_top_level_names(items: Any) -> Optional[set[str]]:
    if not isinstance(items, list):
        return None
    out: set[str] = set()
    for x in items:
        t = top_level_import_name(x)
        if t:
            out.add(t)
    return out


def compute_agent_wall_time_sec(job_summary: Dict[str, Any]) -> Optional[float]:
    ts = job_summary.get("timestamps") or {}
    a0 = ts.get("agent_start")
    a1 = ts.get("agent_end")
    if isinstance(a0, (int, float)) and isinstance(a1, (int, float)) and a1 >= a0:
        return round(float(a1 - a0), 3)
    return None


def derive_c2_cuda_status(cuda_res: Optional[Dict[str, Any]]) -> str:
    if not cuda_res:
        return "missing"
    status = normalize_status(cuda_res.get("status"))
    # Prefer observed.cuda_available (new scripts) then metrics.cuda_available (legacy).
    cuda_available = safe_get(cuda_res, "observed", "cuda_available", default=None)
    if cuda_available is None:
        cuda_available = safe_get(cuda_res, "metrics", "cuda_available", default=None)
    # rule: cuda_available must be True to count as success (if value is present)
    if status == "success" and (cuda_available is True or cuda_available is None):
        return "success"
    if status == "skipped":
        return "skipped"
    # if stage ran but cuda_available false -> failed
    if cuda_available is False:
        return "failed"
    return status or "unknown"


def read_env_size_mb(env_res: Optional[Dict[str, Any]]) -> Optional[float]:
    if not env_res:
        return None
    # New scripts: observed.env_prefix_size_MB (note the capital MB)
    v = safe_get(env_res, "observed", "env_prefix_size_MB", default=None)
    f = coerce_float(v)
    if f is not None:
        return round(f, 3)

    # Legacy fallback: metrics.*
    for k in ("env_prefix_size_mb", "env_size_mb", "prefix_size_mb", "env_prefix_size_MB"):
        v2 = safe_get(env_res, "metrics", k, default=None)
        if v2 is None:
            v2 = env_res.get(k)
        f2 = coerce_float(v2)
        if f2 is not None:
            return round(f2, 3)
    return None

def read_hallucination_counts(h_res: Optional[Dict[str, Any]]) -> Dict[str, int]:
    # default 0 (so sums work)
    out: Dict[str, int] = {"path": 0, "version": 0, "capability": 0}
    if not h_res:
        return out

    # support multiple possible schemas
    out["path"] = coerce_int(safe_get(h_res, "hallucinations", "path", "count")) or 0
    out["version"] = coerce_int(safe_get(h_res, "hallucinations", "version", "count")) or 0
    out["capability"] = coerce_int(safe_get(h_res, "hallucinations", "capability", "count")) or 0

    # alternative flat keys
    if out["path"] == 0:
        out["path"] = coerce_int(safe_get(h_res, "metrics", "path_hallucinations_count")) or out["path"]
    if out["version"] == 0:
        out["version"] = coerce_int(safe_get(h_res, "metrics", "version_hallucinations_count")) or out["version"]
    if out["capability"] == 0:
        out["capability"] = coerce_int(safe_get(h_res, "metrics", "capability_hallucinations_count")) or out["capability"]

    return out


def _pick_baseline(job_summary: Dict[str, Any]) -> str:
    # Backward compatible: old job_summary used baseline_backend/baseline_input
    return (
        (job_summary.get("baseline") or "").strip()
        or (job_summary.get("baseline_backend") or "").strip()
        or (job_summary.get("baseline_input") or "").strip()
    )


def build_row(job_dir: Path) -> Tuple[Dict[str, Any], List[str]]:
    problems: List[str] = []

    job_summary = read_json(job_dir / "job_summary.json") or {}
    run_id = job_summary.get("run_id", "")
    job_id = job_summary.get("job_id", "")
    repo_url = job_summary.get("repo_url", "")
    commit_sha = job_summary.get("commit_sha", "")
    baseline = _pick_baseline(job_summary)
    hw_bucket = (job_summary.get("hardware_bucket") or "").strip()
    script_id = str(job_summary.get("script_id") or "").strip()

    repo_full_name = parse_repo_full_name(repo_url)

    # paths
    docker_dir = job_dir / "docker"
    agent_dir = job_dir / "agent"
    bench_dir = job_dir / "benchmark"
    build_output_dir = bench_dir / "build_output"


    # docker info
    image_id = try_read_text(docker_dir / "image_id.txt").strip()
    nvsmi = try_read_text(docker_dir / "nvidia_smi.txt")
    gpu_name, driver_ver, cuda_rt = parse_nvidia_smi_text(nvsmi)

    # stage results
    pyright_res = load_stage_results(build_output_dir, "pyright")
    cpu_res = load_stage_results(build_output_dir, "cpu")
    cuda_res = load_stage_results(build_output_dir, "cuda")
    s1_res = load_stage_results(build_output_dir, "single_gpu")
    s2_res = load_stage_results(build_output_dir, "multi_gpu")
    hall_res = load_stage_results(build_output_dir, "hallucination")
    env_res = load_stage_results(build_output_dir, "env_size")

    # C0 (pyright missing imports)
    c0_missing: Optional[int] = None
    c0_total: Optional[int] = None
    c0_total_baseline: Optional[int] = None

    if pyright_res:
        # New per-repo scripts write top-level counts.
        c0_missing = coerce_int(pyright_res.get("missing_packages_count"))
        if c0_missing is None:
            c0_missing = coerce_int(pyright_res.get("missing_imports_count"))

        c0_total = coerce_int(pyright_res.get("total_imported_packages_count"))
        if c0_total is None:
            c0_total = coerce_int(pyright_res.get("total_imports_count"))

        # Legacy fallbacks (older harness schemas)
        if c0_missing is None:
            c0_missing = coerce_int(safe_get(pyright_res, "metrics", "missing_packages_count"))
            if c0_missing is None:
                c0_missing = coerce_int(safe_get(pyright_res, "metrics", "missing_imports_count"))
            if c0_missing is None:
                c0_missing = coerce_int(safe_get(pyright_res, "meta", "metrics", "missing_packages_count"))
            if c0_missing is None:
                c0_missing = coerce_int(safe_get(pyright_res, "meta", "metrics", "missing_imports_count"))

        if c0_total is None:
            for k in (
                "total_imported_packages_count",
                "total_imports_count",
                "total_imports",
                "imports_total",
                "total_imported_modules_count",
                "total_imported_packages",
            ):
                c0_total = coerce_int(safe_get(pyright_res, "metrics", k))
                if c0_total is not None:
                    break
            if c0_total is None:
                for k in (
                    "total_imported_packages_count",
                    "total_imports_count",
                    "total_imports",
                    "imports_total",
                    "total_imported_modules_count",
                    "total_imported_packages",
                ):
                    c0_total = coerce_int(safe_get(pyright_res, "meta", "metrics", k))
                    if c0_total is not None:
                        break

    # fallback: if benchmark wrote a helper file (e.g., compute_total_imports.py), read it
    if c0_total is None:
        imports_path = build_output_dir / "pyright" / "imports_detected.json"
        imports_res = read_json(imports_path)
        if imports_res:
            c0_total = (
                coerce_int(safe_get(imports_res, "metrics", "total_imported_packages_count"))
                or coerce_int(imports_res.get("total_imported_packages_count"))
                or c0_total
            )

    # Preferred stable denominator: baseline imports computed from git-tracked files.
    baseline_path = build_output_dir / "pyright" / "baseline_imports.json"
    baseline_res = read_json(baseline_path)
    if isinstance(baseline_res, dict):
        c0_total_baseline = (
            coerce_int(baseline_res.get("total_imported_packages_count"))
            or coerce_int(safe_get(baseline_res, "metrics", "total_imported_packages_count"))
            or coerce_int(safe_get(baseline_res, "meta", "metrics", "total_imported_packages_count"))
        )

    # Guardrail: if we ended up with missing > total, cap and record a problem.
    if (
        c0_total is not None
        and c0_total > 0
        and c0_missing is not None
        and c0_missing > c0_total
    ):
        problems.append(f"c0_missing_gt_total (capped): missing={c0_missing} total={c0_total}")
        c0_missing = c0_total


    # C1/C3/C4 (repo-specific entrypoints)
    c1 = normalize_status(cpu_res.get("status")) if cpu_res else "missing"
    c3 = normalize_status(s1_res.get("status")) if s1_res else "missing"
    c4 = normalize_status(s2_res.get("status")) if s2_res else "missing"

    # C2
    c2 = derive_c2_cuda_status(cuda_res)

    # C5
    h = read_hallucination_counts(hall_res)

    # agent time
    agent_wall = compute_agent_wall_time_sec(job_summary)

    # env size
    env_mb = read_env_size_mb(env_res)

    row = {
        "run_id": run_id,
        "job_id": job_id,
        "repo_full_name": repo_full_name,
        "repo_url": repo_url,
        "commit_sha": commit_sha,
        "baseline": baseline,
        "hardware_bucket": hw_bucket,
        "script_id": script_id,
        "supports_cpu": "",
        "supports_single_gpu": "",
        "supports_multi_gpu": "",
        "c0_missing_imports": c0_missing if c0_missing is not None else "",
        "c0_total_imports": c0_total if c0_total is not None else "",
        "c0_total_imports_baseline": c0_total_baseline if c0_total_baseline is not None else "",
        "c1_cpu_status": c1,
        "c2_cuda_status": c2,
        "c3_single_gpu_status": c3,
        "c4_multi_gpu_status": c4,
        "c5_path_hallucinations_count": h["path"],
        "c5_version_hallucinations_count": h["version"],
        "c5_capability_hallucinations_count": h["capability"],
        "agent_wall_time_sec": agent_wall if agent_wall is not None else "",
        "env_prefix_size_mb": env_mb if env_mb is not None else "",
        "image_id": image_id,
        "gpu_name": gpu_name,
        "driver_version": driver_ver,
        "cuda_runtime": cuda_rt,
        "job_dir": str(job_dir),
    }

    # basic missing diagnostics
    if not build_output_dir.exists():
        problems.append("missing benchmark/build_output")
    if not agent_dir.exists():
        problems.append("missing agent dir")
    if not docker_dir.exists():
        problems.append("missing docker dir")

    return row, problems


def write_csv(path: Path, rows: List[Dict[str, Any]], columns: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in columns})


def _count_statuses(statuses: List[str]) -> Tuple[int, int, int]:
    """Return (success, failed, skipped). Unknown/missing are counted as failed."""
    suc = fail = skip = 0
    for s in statuses:
        ns = normalize_status(s)
        if ns == "success":
            suc += 1
        elif ns == "skipped":
            skip += 1
        else:
            fail += 1
    return suc, fail, skip


def build_summary_rows(job_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # group by (run_id, baseline)
    groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in job_rows:
        rid = str(r.get("run_id", "") or "")
        b = str(r.get("baseline", "") or "")
        groups[(rid, b)].append(r)

    out: List[Dict[str, Any]] = []

    for (run_id, baseline), rows in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
        job_count = len(rows)

        # C0 sums
        c0_missing_sum = sum(coerce_int(r.get("c0_missing_imports")) or 0 for r in rows)
        c0_total_sum_reported = sum(coerce_int(r.get("c0_total_imports")) or 0 for r in rows)

        def c0_total_for_row(r: Dict[str, Any]) -> int:
            repo_full_name = str(r.get("repo_full_name") or "").strip()
            if not repo_full_name:
                repo_full_name = parse_repo_full_name(str(r.get("repo_url") or ""))

            repo_ref_total = c0_reference_total_for_repo(repo_full_name)
            if repo_ref_total is not None and repo_ref_total > 0:
                return repo_ref_total

            baseline_total = coerce_int(r.get("c0_total_imports_baseline"))
            if baseline_total is not None and baseline_total > 0:
                return baseline_total
            return coerce_int(r.get("c0_total_imports")) or 0

        c0_total_sum = sum(c0_total_for_row(r) for r in rows)
        c0_missing_over_total = f"{c0_missing_sum}/{c0_total_sum}" if c0_total_sum > 0 else ""

        # C1~C4
        cpu_rows = [r for r in rows if _is_supported(r.get("supports_cpu"))]
        single_rows = [r for r in rows if _is_supported(r.get("supports_single_gpu"))]
        multi_rows = [r for r in rows if _is_supported(r.get("supports_multi_gpu"))]

        c1_s, c1_f, c1_k = _count_statuses([str(r.get("c1_cpu_status") or "") for r in cpu_rows])
        c2_s, c2_f, c2_k = _count_statuses([str(r.get("c2_cuda_status") or "") for r in rows])
        c3_s, c3_f, c3_k = _count_statuses([str(r.get("c3_single_gpu_status") or "") for r in single_rows])
        c4_s, c4_f, c4_k = _count_statuses([str(r.get("c4_multi_gpu_status") or "") for r in multi_rows])

        c1_denom = len(cpu_rows)
        c2_denom = job_count
        c3_denom = len(single_rows)
        c4_denom = len(multi_rows)

        # C5
        c5_path_sum = sum(coerce_int(r.get("c5_path_hallucinations_count")) or 0 for r in rows)
        c5_version_sum = sum(coerce_int(r.get("c5_version_hallucinations_count")) or 0 for r in rows)
        c5_cap_sum = sum(coerce_int(r.get("c5_capability_hallucinations_count")) or 0 for r in rows)
        c5_total_sum = c5_path_sum + c5_version_sum + c5_cap_sum

        # agent time avg
        time_vals = [coerce_float(r.get("agent_wall_time_sec")) for r in rows]
        time_vals = [v for v in time_vals if v is not None]
        time_valid = len(time_vals)
        time_avg = round(sum(time_vals) / time_valid, 2) if time_valid > 0 else ""

        # env size avg (GB)
        env_vals_mb = [coerce_float(r.get("env_prefix_size_mb")) for r in rows]
        env_vals_mb = [v for v in env_vals_mb if v is not None]
        env_valid = len(env_vals_mb)
        env_avg_gb = round((sum(env_vals_mb) / env_valid) / 1024.0, 3) if env_valid > 0 else ""

        # Langfuse token metrics (run-level). Exported once per run.
        lf_input = ""
        lf_output = ""
        lf_total = ""
        lf_total_avg = ""
        try:
            any_job_dir = rows[0].get("job_dir") if rows else None
            if any_job_dir:
                run_dir = Path(any_job_dir).resolve().parent.parent
                lf = load_langfuse_run_token_sums(run_dir)
                if lf.get("input_tokens") is not None:
                    lf_input = lf["input_tokens"]
                if lf.get("output_tokens") is not None:
                    lf_output = lf["output_tokens"]
                if lf.get("total_tokens") is not None:
                    lf_total = lf["total_tokens"]
                if isinstance(lf_total, int) and time_valid > 0:
                    lf_total_avg = round(lf_total / time_valid, 3)
        except Exception:
            pass

        def ratio(success: int, denom: int) -> str:
            return f"{success}/{denom}" if denom > 0 else ""

        out.append(
            {
                "run_id": run_id,
                "baseline": baseline,
                "job_count": job_count,
                "c0_missing_sum": c0_missing_sum,
                "c0_total_sum": c0_total_sum,
                "c0_total_sum_reported": c0_total_sum_reported,
                "c0_missing_over_total": c0_missing_over_total,
                "c1_cpu_denom": c1_denom,
                "c2_cuda_denom": c2_denom,
                "c3_single_gpu_denom": c3_denom,
                "c4_multi_gpu_denom": c4_denom,
                "c1_cpu_success_over_all": ratio(c1_s, c1_denom),
                "c2_cuda_success_over_all": ratio(c2_s, c2_denom),
                "c3_single_gpu_success_over_all": ratio(c3_s, c3_denom),
                "c4_multi_gpu_success_over_all": ratio(c4_s, c4_denom),
                "c1_cpu_success": c1_s,
                "c1_cpu_failed": c1_f,
                "c1_cpu_skipped": c1_k,
                "c2_cuda_success": c2_s,
                "c2_cuda_failed": c2_f,
                "c2_cuda_skipped": c2_k,
                "c3_single_gpu_success": c3_s,
                "c3_single_gpu_failed": c3_f,
                "c3_single_gpu_skipped": c3_k,
                "c4_multi_gpu_success": c4_s,
                "c4_multi_gpu_failed": c4_f,
                "c4_multi_gpu_skipped": c4_k,
                "c5_path_sum": c5_path_sum,
                "c5_version_sum": c5_version_sum,
                "c5_capability_sum": c5_cap_sum,
                "c5_total_sum": c5_total_sum,
                "agent_wall_time_avg_sec": time_avg,
                "agent_time_valid_jobs": time_valid,
                "env_size_avg_gb": env_avg_gb,
                "env_valid_jobs": env_valid,
                "agent_tokens_input_sum": lf_input,
                "agent_tokens_output_sum": lf_output,
                "agent_tokens_total_sum": lf_total,
                "agent_tokens_total_avg": lf_total_avg,
            }
        )

    return out


def write_xlsx(path: Path, job_rows: List[Dict[str, Any]], summary_rows: List[Dict[str, Any]]) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed; cannot write xlsx")

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1: jobs
    ws1 = wb.active
    ws1.title = "master_table"
    ws1.append(JOB_COLUMNS)
    for r in job_rows:
        ws1.append([r.get(k, "") for k in JOB_COLUMNS])
    for i, col in enumerate(JOB_COLUMNS, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 12), 60)

    # Sheet 2: summary
    ws2 = wb.create_sheet("summary")
    ws2.append(SUMMARY_COLUMNS)
    for r in summary_rows:
        ws2.append([r.get(k, "") for k in SUMMARY_COLUMNS])
    for i, col in enumerate(SUMMARY_COLUMNS, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 12), 40)

    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="M5: build master tables from job output dirs")
    ap.add_argument("--results-root", required=True, type=Path, help="Directory that contains many job dirs")
    ap.add_argument("--out-dir", default=None, type=Path, help="Output directory (default: <results-root>/_m5)")
    ap.add_argument("--no-xlsx", action="store_true", help="Only write CSV")
    args = ap.parse_args()

    results_root: Path = args.results_root.resolve()
    out_dir: Path = (args.out_dir.resolve() if args.out_dir else (results_root / "_m5").resolve())
    out_dir.mkdir(parents=True, exist_ok=True)

    job_dirs = find_job_dirs(results_root)

    job_rows: List[Dict[str, Any]] = []
    build_log: Dict[str, Any] = {
        "results_root": str(results_root),
        "job_count": len(job_dirs),
        "problems": {},
    }

    categories_idx: Optional[Dict[str, Dict[str, Dict[str, str]]]] = None
    categories_path = resolve_categories_csv(out_dir)
    if categories_path:
        try:
            categories_idx = load_categories_index(categories_path)
            build_log["categories_csv"] = str(categories_path)
        except Exception as e:
            build_log["categories_csv_error"] = str(e)
            categories_idx = None

    for jd in job_dirs:
        row, probs = build_row(jd)
        job_rows.append(row)
        if probs:
            build_log["problems"][str(jd)] = probs

    # Fill applicability flags from scripts_repos_test_categories.csv (if available).
    if categories_idx:
        missing_rows = 0
        for r in job_rows:
            crow = find_categories_row(
                categories_idx,
                script_id=str(r.get("script_id") or ""),
                repo_url=str(r.get("repo_url") or ""),
                repo_full_name=str(r.get("repo_full_name") or ""),
            )
            if not crow:
                missing_rows += 1
                continue
            r["supports_cpu"] = crow.get("supports_cpu", "")
            r["supports_single_gpu"] = crow.get("supports_single_gpu", "")
            r["supports_multi_gpu"] = crow.get("supports_multi_gpu", "")
        if missing_rows:
            build_log["categories_missing_job_rows"] = missing_rows

    # sort job rows for stable output
    job_rows = sorted(job_rows, key=lambda r: (str(r.get("repo_full_name", "")), str(r.get("baseline", "")), str(r.get("job_id", ""))))

    summary_rows = build_summary_rows(job_rows)

    # write CSVs
    csv_jobs = out_dir / "master_table.csv"
    csv_summary = out_dir / "master_summary.csv"
    write_csv(csv_jobs, job_rows, JOB_COLUMNS)
    write_csv(csv_summary, summary_rows, SUMMARY_COLUMNS)

    # write XLSX
    xlsx_path = out_dir / "master_table.xlsx"
    if not args.no_xlsx:
        try:
            write_xlsx(xlsx_path, job_rows, summary_rows)
        except Exception as e:
            build_log["xlsx_error"] = str(e)

    write_json(out_dir / "m5_build_log.json", build_log)

    print(f"[M5] Wrote: {csv_jobs}")
    print(f"[M5] Wrote: {csv_summary}")
    if not args.no_xlsx:
        print(f"[M5] Wrote: {xlsx_path}")
    print(f"[M5] Log:   {out_dir / 'm5_build_log.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
